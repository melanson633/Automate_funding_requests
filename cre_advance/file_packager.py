from __future__ import annotations

"""Package funding request outputs.

This module aligns normalized Driver rows with segmented invoice pages,
building the final Excel workbook, reordered PDF, and a JSON report.
"""

import json
from pathlib import Path
from typing import List

import pandas as pd
from Levenshtein import ratio
from openpyxl import load_workbook
from pypdf import PdfReader, PdfWriter

from .utils.logging import get_logger

logger = get_logger(__name__)


def _detect_duplicates(df: pd.DataFrame, manifest: List[dict]) -> dict:
    """Return duplicate invoice numbers in Excel rows and PDF manifest."""
    excel_counts = df.get("invoice_number").astype(str).str.strip().value_counts()
    pdf_counts = pd.Series(
        [m.get("invoice_number", "").strip() for m in manifest]
    ).value_counts()
    dup_excel = excel_counts[excel_counts > 1].index.tolist()
    dup_pdf = pdf_counts[pdf_counts > 1].index.tolist()
    return {"excel": dup_excel, "pdf": dup_pdf}


def _match_invoices(
    df: pd.DataFrame, manifest: List[dict], cfg: dict | None = None
) -> tuple[List[dict], list[int], list[int]]:
    """Return matched invoices and lists of unmatched row and PDF indices."""
    cfg = cfg or {}
    pkg_cfg = cfg.get("packager", {})
    v_ratio_thres = float(pkg_cfg.get("vendor_ratio_threshold", 0.8))
    amt_tol = float(pkg_cfg.get("amount_tolerance", 0.01))
    score_thres = float(pkg_cfg.get("score_threshold", 2.0))

    used: set[int] = set()
    ordered: List[dict] = []
    unmatched_rows: list[int] = []

    for idx, row in df.iterrows():
        row_num = str(row.get("invoice_number", "")).strip()
        row_vendor = str(row.get("vendor", "")).strip().lower()
        row_amt = float(row.get("amount", 0))
        row_date = (
            str(pd.to_datetime(row.get("date")).date()) if row.get("date") else ""
        )

        best = None
        best_score = 0.0
        for m_idx, item in enumerate(manifest):
            if m_idx in used:
                continue
            inv_num = str(item.get("invoice_number", "")).strip()
            inv_vendor = str(item.get("vendor", "")).strip().lower()
            inv_amt = float(item.get("amount", 0))
            inv_date = (
                str(pd.to_datetime(item.get("date")).date()) if item.get("date") else ""
            )

            score = 0.0
            if row_num and inv_num and row_num == inv_num:
                score += 3.0
            if row_vendor and inv_vendor:
                v_ratio = ratio(row_vendor, inv_vendor)
                if v_ratio >= v_ratio_thres:
                    score += v_ratio
            if abs(row_amt - inv_amt) <= amt_tol:
                score += 1.0
            if row_date and inv_date and row_date == inv_date:
                score += 1.0

            if score > best_score:
                best_score = score
                best = (m_idx, item)

        if best is not None and best_score >= score_thres:
            used.add(best[0])
            ordered.append(best[1])
        else:
            unmatched_rows.append(idx)

    unmatched_pdf = [i for i in range(len(manifest)) if i not in used]
    return ordered, unmatched_rows, unmatched_pdf


def _build_pdf(manifest: List[dict], pdf_path: Path, dest: Path) -> None:
    """Reorder pages from ``pdf_path`` according to ``manifest``."""
    reader = PdfReader(str(pdf_path))
    writer = PdfWriter()
    for item in manifest:
        start = int(item.get("start_page", 1)) - 1
        end = int(item.get("end_page", start + 1)) - 1
        for page in reader.pages[start : end + 1]:
            writer.add_page(page)
    with dest.open("wb") as f:
        writer.write(f)


def _write_excel(
    driver_df: pd.DataFrame,
    invoice_df: pd.DataFrame,
    template: Path,
    dest: Path,
    status: List[str] | None = None,
) -> None:
    """Create workbook with hidden Driver and Invoice Log and match status."""
    wb = load_workbook(template)
    if "Driver" not in wb.sheetnames:
        wb.create_sheet("Driver")
    if "Invoice Log" in wb.sheetnames:
        del wb["Invoice Log"]
    wb.create_sheet("Invoice Log")
    if status is not None:
        if "Match Status" in wb.sheetnames:
            del wb["Match Status"]
        wb.create_sheet("Match Status")

    for row in driver_df.itertuples(index=False, name=None):
        wb["Driver"].append(list(row))
    wb["Driver"].sheet_state = "hidden"

    for row in invoice_df.itertuples(index=False, name=None):
        wb["Invoice Log"].append(list(row))

    if "amount" in invoice_df.columns:
        total = invoice_df["amount"].sum()
        wb["Invoice Log"].append(["Total", total])

    if status is not None:
        for s in status:
            wb["Match Status"].append([s])
        wb["Match Status"].sheet_state = "hidden"

    wb.save(dest)


def package(
    normalized_df: pd.DataFrame,
    manifest: List[dict],
    template_path: str | Path,
    pdf_path: str | Path,
    output_dir: str | Path,
    cfg: dict | None = None,
    metrics: dict | None = None,
) -> dict:
    """Build funding package and return paths to artefacts."""
    cfg = cfg or {}
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    logger.info("Packaging deliverables", extra={"context": "package"})
    ordered, unmatched_rows, unmatched_pdf = _match_invoices(
        normalized_df, manifest, cfg
    )
    unmatched_threshold = float(cfg.get("unmatched_threshold", 0.4))
    warning = None
    if (
        len(normalized_df.index)
        and len(unmatched_rows) / len(normalized_df.index) > unmatched_threshold
    ):
        warning = "High unmatched ratio"
        logger.warning(warning, extra={"context": "package"})

    ordered_df = normalized_df.drop(index=unmatched_rows).reset_index(drop=True)

    if metrics is not None:
        metrics["rows_processed"] = len(normalized_df.index)
        metrics["rows_unmatched"] = len(unmatched_rows)
        metrics["pdf_invoices"] = len(manifest)
        metrics["pdf_unmatched"] = len(unmatched_pdf)

    pdf_out = output_dir / "invoices.pdf"
    _build_pdf(ordered, Path(pdf_path), pdf_out)

    excel_out = output_dir / "request.xlsx"
    status = [
        "matched" if idx not in unmatched_rows else "unmatched"
        for idx in range(len(normalized_df.index))
    ]
    _write_excel(normalized_df, ordered_df, Path(template_path), excel_out, status)

    dups = _detect_duplicates(normalized_df, manifest)

    report = {
        "total_rows": len(normalized_df.index),
        "unmatched_rows": [
            {
                "index": idx,
                "invoice_number": normalized_df.loc[idx, "invoice_number"],
                "vendor": normalized_df.loc[idx, "vendor"],
                "amount": normalized_df.loc[idx, "amount"],
            }
            for idx in unmatched_rows
        ],
        "unmatched_pdf": [manifest[i] for i in unmatched_pdf],
        "output_excel": str(excel_out),
        "output_pdf": str(pdf_out),
        "duplicate_excel_invoices": dups["excel"],
        "duplicate_pdf_invoices": dups["pdf"],
        "warning": warning,
        "metrics": metrics or {},
    }
    report_out = output_dir / "report.json"
    report_out.write_text(json.dumps(report, indent=2))

    logger.info("Finished packaging", extra={"context": "package"})

    return {
        "excel": str(excel_out),
        "pdf": str(pdf_out),
        "report": str(report_out),
        "unmatched_rows": unmatched_rows,
        "unmatched_pdf": unmatched_pdf,
        "duplicate_excel_invoices": dups["excel"],
        "duplicate_pdf_invoices": dups["pdf"],
        "warning": warning,
        "metrics": metrics or {},
    }
