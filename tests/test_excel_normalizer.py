from __future__ import annotations

import sys
import tempfile
import types
from pathlib import Path
from typing import Any

import pandas as pd
import pytest
import yaml
from openpyxl import Workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))  # noqa: E402

# Mock the genai module structure
google_mock = types.ModuleType("google")
genai_mock = types.ModuleType("genai")
genai_types_mock = types.ModuleType("types")
genai_mock.types = genai_types_mock
google_mock.genai = genai_mock

sys.modules.setdefault("google", google_mock)
sys.modules.setdefault("google.genai", genai_mock)
sys.modules.setdefault("google.genai.types", genai_types_mock)
from cre_advance import excel_normalizer  # noqa: E402


def _create_workbook(path: Path, header_row: int = 4) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Driver"
    for _ in range(header_row - 1):
        ws.append([])
    ws.append(["Date", "Amount"])
    ws.append(["2024-01-01", "100"])
    ws.append(["2024-02-01", "(200)"])
    wb.save(path)


def test_normalize_basic(monkeypatch) -> None:
    with tempfile.TemporaryDirectory() as tmp:
        excel_path = Path(tmp) / "test.xlsx"
        _create_workbook(excel_path)

        def fake_map(headers, samples, fields):
            return {"Date": "Date", "Amount": "Amount"}

        def fake_build(headers, samples):
            return {}

        monkeypatch.setattr(
            excel_normalizer.ai_gemini,
            "map_headers",
            fake_map,
        )
        monkeypatch.setattr(
            excel_normalizer.ai_gemini,
            "build_schema",
            fake_build,
        )

        cfg = {
            "lender": "example_lender",
            "excel": {"fields": ["Date", "Amount"]},
        }
        normalized, raw = excel_normalizer.normalize([excel_path], cfg)

        assert list(normalized.columns) == ["Date", "Amount"]
        assert pd.api.types.is_datetime64_any_dtype(normalized["Date"])
        assert normalized.loc[1, "Amount"] == -200.0
        assert not raw.empty


def test_header_row_config(monkeypatch) -> None:
    with tempfile.TemporaryDirectory() as tmp:
        excel_path = Path(tmp) / "hdr.xlsx"
        _create_workbook(excel_path, header_row=3)

        monkeypatch.setattr(
            excel_normalizer.ai_gemini,
            "map_headers",
            lambda h, s, f: {"Date": "Date", "Amount": "Amount"},
        )
        monkeypatch.setattr(excel_normalizer.ai_gemini, "build_schema", lambda h, s: {})

        cfg = {"lender": "x", "excel": {"fields": ["Date", "Amount"], "header_row": 3}}
        normalized, _ = excel_normalizer.normalize([excel_path], cfg)

        assert list(normalized.columns) == ["Date", "Amount"]


def test_auto_save_schema(monkeypatch, tmp_path) -> None:
    excel_path = tmp_path / "auto.xlsx"
    _create_workbook(excel_path)

    monkeypatch.setattr(
        excel_normalizer.ai_gemini,
        "map_headers",
        lambda h, s, f: {},
    )

    monkeypatch.setattr(
        excel_normalizer.ai_gemini,
        "build_schema",
        lambda h, s: {
            "mapping": {"Date": "InvoiceDate"},
            "fields": ["InvoiceDate"],
        },
    )

    cfg = {"lender": "example_lender", "excel": {"fields": ["InvoiceDate"]}}

    version_dir = Path(__file__).resolve().parents[1] / "configs" / "schema_versions"
    if version_dir.exists():
        for f in version_dir.glob("example_lender_*.yaml"):
            f.unlink()

    normalized, _ = excel_normalizer.normalize([excel_path], cfg)

    files = list(version_dir.glob("example_lender_*.yaml"))
    assert len(files) == 1
    assert "InvoiceDate" in normalized.columns
    with files[0].open() as f:
        data = yaml.safe_load(f)
    assert data["excel"]["mapping"]["Date"] == "InvoiceDate"
    files[0].unlink()


def test_fuzzy_fallback(monkeypatch) -> None:
    with tempfile.TemporaryDirectory() as tmp:
        excel_path = Path(tmp) / "fuzzy.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Driver"
        ws.append([])
        ws.append([])
        ws.append([])
        ws.append(["Date", "Amt"])
        ws.append(["2024-01-01", "100"])
        wb.save(excel_path)

        monkeypatch.setattr(
            excel_normalizer.ai_gemini,
            "map_headers",
            lambda h, s, f: {},
        )
        monkeypatch.setattr(excel_normalizer.ai_gemini, "build_schema", lambda h, s: {})

        cfg = {
            "lender": "x",
            "excel": {"fields": ["Date", "Amount"], "fuzzy_ratio": 0.5},
        }

        normalized, _ = excel_normalizer.normalize([excel_path], cfg)

        assert list(normalized.columns) == ["Date", "Amount"]


def test_read_workbook_missing_sheet(tmp_path: Path) -> None:
    path = tmp_path / "wb.xlsx"
    _create_workbook(path)
    with pytest.raises(ValueError):
        excel_normalizer._read_workbook(path, "Missing", 4)


def test_read_workbook_header_fallback(tmp_path: Path) -> None:
    path = tmp_path / "wb.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append([])
    ws.append(["A", "B"])
    ws.append([1, 2])
    wb.save(path)

    df = excel_normalizer._read_workbook(path, "Data", 1)
    assert list(df.columns) == ["A", "B"]


def test_normalize_report_detection(monkeypatch) -> None:
    with tempfile.TemporaryDirectory() as tmp:
        gl_path = Path(tmp) / "gl.xlsx"
        ed_path = Path(tmp) / "ed.xlsx"

        # General ledger workbook
        wb1 = Workbook()
        ws1 = wb1.active
        ws1.title = "Report1"
        for _ in range(5):
            ws1.append([])
        ws1.append(["Date", "Amount"])
        ws1.append(["2024-01-01", "100"])
        wb1.save(gl_path)

        # Expense distribution workbook
        wb2 = Workbook()
        ws2 = wb2.active
        ws2.title = "Expense Distribution Report"
        for _ in range(2):
            ws2.append([])
        ws2.append(["Date", "Amount"])
        ws2.append(["2024-01-02", "200"])
        wb2.save(ed_path)

        monkeypatch.setattr(
            excel_normalizer.ai_gemini,
            "map_headers",
            lambda h, s, f: {"Date": "Date", "Amount": "Amount"},
        )
        monkeypatch.setattr(excel_normalizer.ai_gemini, "build_schema", lambda h, s: {})

        cfg = {"lender": "x", "excel": {"fields": ["Date", "Amount"]}}
        metrics: dict[str, Any] = {}
        normalized, _ = excel_normalizer.normalize([gl_path, ed_path], cfg, metrics)

        assert normalized.shape[0] == 2
        types = {m["type"] for m in metrics["report_detections"]}
        assert "general_ledger" in types
        assert "expense_distribution" in types
        assert all("method" in m and "confidence" in m for m in metrics["report_detections"])
        assert all(m["method"] == "heuristic" for m in metrics["report_detections"])
