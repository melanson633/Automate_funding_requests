from __future__ import annotations

import tempfile
import tempfile
from pathlib import Path

import pytest
from openpyxl import Workbook

from cre_advance import excel_normalizer


@pytest.mark.parametrize(
    "sheet_name, header_row, expected",
    [
        ("Report1", 6, "general_ledger"),
        ("Expense Distribution Report", 3, "expense_distribution"),
        ("DRIVER", 4, "funding_template"),
    ],
)
def test_detect_report_type_known(
    sheet_name: str, header_row: int, expected: str
) -> None:
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "wb.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        for _ in range(header_row - 1):
            ws.append([])
        ws.append(["A", "B"])
        wb.save(path)

        result = excel_normalizer.detect_report_type(path)
        assert result == {
            "type": expected,
            "sheet_name": sheet_name,
            "header_row": header_row,
            "confidence": 1.0,
            "method": "heuristic",
        }


def test_detect_report_type_unknown(tmp_path: Path) -> None:
    path = tmp_path / "unknown.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "MySheet"
    ws.append([])
    ws.append(["Col1", "Col2"])
    ws.append([1, 2])
    wb.save(path)

    result = excel_normalizer.detect_report_type(path)
    assert result["type"] == "unknown"
    assert result["sheet_name"] == "MySheet"
    assert result["header_row"] == 2
    assert result["method"] == "heuristic"
    assert result["confidence"] > 0


def test_detect_report_type_no_confidence(tmp_path: Path) -> None:
    path = tmp_path / "blank.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    wb.save(path)

    result = excel_normalizer.detect_report_type(path, {"use_ai_detection": False})
    assert result == {
        "type": "unknown",
        "sheet_name": None,
        "header_row": None,
        "confidence": 0.0,
        "method": "heuristic",
    }


def test_detect_report_type_ai_fallback(tmp_path: Path, monkeypatch) -> None:
    path = tmp_path / "ai.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    wb.save(path)

    calls = {"count": 0}

    def fake_detect(file_path):
        calls["count"] += 1
        return {"sheet_name": "Sheet1", "header_row": 1, "confidence": 0.9}

    monkeypatch.setattr(excel_normalizer, "_ai_detect", fake_detect)
    result = excel_normalizer.detect_report_type(
        path, {"use_ai_detection": True, "ai_detection_threshold": 0.8}
    )

    assert result["method"] == "ai"
    assert result["sheet_name"] == "Sheet1"
    assert result["header_row"] == 1
    assert calls["count"] == 1


def test_detect_report_type_ai_cached(tmp_path: Path, monkeypatch) -> None:
    path = tmp_path / "cache.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    wb.save(path)

    calls = {"count": 0}

    def fake_detect(file_path: Path):
        calls["count"] += 1
        return {"sheet_name": "Sheet1", "header_row": 1, "confidence": 0.9}

    monkeypatch.setattr(excel_normalizer.ai_gemini, "detect_excel_structure", fake_detect)
    excel_normalizer._ai_detect.cache_clear()  # type: ignore[attr-defined]

    excel_normalizer.detect_report_type(path, {"force_ai_detection": True})
    excel_normalizer.detect_report_type(path, {"force_ai_detection": True})

    assert calls["count"] == 1
